import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import os
import pandas as pd
import random as rand
import numpy as np

debug=True
# pdebug=True

class colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

#?-------Style functions------------------
def setTableDimensions(writingSheet, columnWidth, rowHeight):
    startColumn='A' #go +1 in loop #!  limit to 'G'
    endColumn='G'

    for row in range(0, 17): 
        writingSheet.row_dimensions[row].height=rowHeight

    for column in range(ord(startColumn), ord(endColumn)+1): 
        writingSheet.column_dimensions[chr(column)].width=columnWidth

def centerTableAlignment(writingSheet):
    startColumn='A'
    endColumn='G'

    for row in range(1, 17): #start at row 1, end at 16, increment by 2
        for column in range(ord(startColumn), ord(endColumn)+1):
            writingSheet[chr(column)+str(row)].alignment=Alignment(horizontal='center', vertical='center')   

def setFontStyles(writingSheet, fontSize):
    for row in range(1, 17, 2):
        writingSheet[f'A{row}'].font=Font(size=fontSize, bold=True)

    for column in range(ord('A'), ord('G')+1):
        writingSheet[f'{chr(column)}1'].font=Font(size=fontSize, bold=True)

def setTimeIntervals(writingSheet):
    #populate writeSheet with time intervals
    dayList=["Luni", "Marți", "Miercuri", "Joi", "Vineri", "Sâmbătă"]
    hourList=["8.00-9.30", "9.45-11.15", "11.30-13.00", "13.30-15.00", "15.15-16.45", "17.00-18.30", "18.45-20.15"]

    #populate writeSheet with time intervals
    startRow=3
    for timeInterval in hourList:
        writingSheet.cell(row=startRow, column=1, value=timeInterval)
        startRow+=2        

    #populate writeSheet with weekdays
    startColumn=2
    for day in dayList:
        writingSheet.cell(row=1, column=startColumn, value=day)
        startColumn+=1

def applyDefaultMergeStyle(writingSheet):
    startColumn='A' #go +1 in loop #!  limit to 'G'
    endColumn='G'

    for row in range(1, 17, 2): #start at row 1, end at 16, increment by 2
        for column in range(ord(startColumn), ord(endColumn)+1):
            currentCell=writingSheet[chr(column)+str(row)]
            underCell=writingSheet[chr(column)+str(row+1)]

            # if row!=1 and column!=65:
            #     currentCell.value='#'


            #data in both
            if currentCell.value is None and underCell.value is None:
                writingSheet.merge_cells(f'{currentCell.coordinate}:{underCell.coordinate}')
                currentCell.value='#'
                continue
                        

            #only up data
            if currentCell.value != None and underCell.value is None:
                writingSheet.merge_cells(f'{currentCell.coordinate}:{underCell.coordinate}')
                continue

            #no up data, ## in lower cell
            if currentCell.value is None and underCell.value=='#':
                writingSheet.merge_cells(f'{currentCell.coordinate}:{underCell.coordinate}')
                currentCell.value='#'
                continue

            #data only in lower
            if currentCell.value is None and underCell.value!=None:
                writingSheet.merge_cells(f'{currentCell.coordinate}:{underCell.coordinate}')
                currentCell.value='#'
                continue            
#?----------------------------------------


#?-------Functional functions-------------
def isInMergedRange(schedule, cell):
    if cell.coordinate in schedule.merged_cells:
        return True
    else:
        return False
    
def saveScheduleTable(writingBook, fileName):
    if fileName is None:
        writingBook.close()
        return
    
    if '/' in fileName:
        fileName=fileName.replace('/', '-')

    filePath=f"timetable/teacher/{fileName}.xlsx"

    if os.path.exists(filePath):
        os.remove(filePath)
    
    writingBook.save(filePath)
    writingBook.close()

def getGroupNames(schedule):
    #lista grupe dupa nume
    fcimGroups=[]
    
    dataRow=""
    #get group start row
    for row in range(1, 20):
        #print(f"B{row}=={schedule[f'B{row}'].value}")
        if schedule[f'B{row}'].value =='Grupele':
            dataRow=row
            break
         
    #popularea listei cu numele grupelor din excel
    for cell in schedule.iter_cols(min_row=dataRow, max_row=dataRow, min_col=5):
        if cell[0].value is not None :
            fcimGroups.append(cell[0].value)
        else:
            break

    return fcimGroups, dataRow

def getExcellFilenames():
    excelFiles=[]
    for filename in os.listdir():
        if '.xlsx' in filename and '$' not in filename:
            excelFiles.append(filename)
    
    excelFiles[0],excelFiles[2],excelFiles[3]=excelFiles[3],excelFiles[0], excelFiles[2]
    return excelFiles
        
def getPDFfilenames():
    folder_path='pdfs'
    filePaths=[]

    for filename in os.listdir(folder_path):
        file_path=os.path.join(folder_path, filename)
        filePaths.append(file_path)

    filePaths[0], filePaths[2], filePaths[3]=filePaths[3], filePaths[0], filePaths[2]
    return filePaths
#?----------------------------------------
        

#!-------______--------------------------------------
def extractAllTextFromExcel(file_path):
    df = pd.read_excel(file_path, header=None)
    stacked_series = df.stack()
    all_text = stacked_series.astype(str).tolist()

    return all_text

def countInstancesInText(name, text):
    # return text.count(name) + text.count(name.replace(' ', ''))
    instances=0
    for string in text:
        if (name in string) or (name.replace(' ', '') in string):
            instances+=1

    return instances

def __buildTreeWithPaths(directory_path):
    tree = {}

    for root, dirs, files in os.walk(directory_path):
        current_level = tree
        path = root.split(os.path.sep)

        for i, folder in enumerate(path):
            if folder not in current_level:
                current_level[folder] = {} if i < len(path) - 1 else {'files': []}
            current_level = current_level[folder]

        if files:
            current_level['files'].extend(files)

    return tree

def __getFilePaths(tree, current_path='', file_paths=None):
    if file_paths is None:
        file_paths = []

    for key, value in tree.items():
        new_path = os.path.join(current_path, key) if current_path else key

        if isinstance(value, dict):
            __getFilePaths(value, current_path=new_path, file_paths=file_paths)
        elif key == 'files':
            for file in value:
                file_paths.append(os.path.join(new_path, file))

    return file_paths

def getPathsList():
    directory_path = 'timetable'
    tree = __buildTreeWithPaths(directory_path)
    excel_file_paths = __getFilePaths(tree)

    filePaths=[]
    for path in excel_file_paths:
        if path.endswith('.xlsx') or path.endswith('.xls'):
            if '$' not in path and 'teacher' not in path:
                filePaths.append(path.replace('files/', ''))
    
    return filePaths

def transferProfClasses(name, instances, readTable, insertTable, tableName):
    validRows=[3, 5, 7, 9, 11, 13, 15]
    foundInstances=0
  
    nospName=name.replace(' ', '')
    thrownPoints=np.zeros((16, 7), dtype=bool) 

    while(foundInstances!=instances):
        randRow=validRows[rand.randint(0, 6)]  
        colIndex=rand.randrange(1, 7)
        randCol=chr(ord('A') + colIndex)
        
        if thrownPoints[randRow, colIndex] or (readCellData is None):
            continue

        readCell=readTable[f'{randCol}{randRow}']
        readCellData=readCell.value

        readCellNext=readTable[f'{randCol}{randRow+1}']
        readCellNextData=readCellNext.value

        try:
            if isInMergedRange(readTable, readCell):
                # writingSheet.merge_cells(f'{currentCell.coordinate}:{underCell.coordinate}')
                insertTable[f'{randCol}{randRow}'].value=readCellData+tableName
                insertTable[f'{randCol}{randRow+1}'].value=readCellData+tableName

            if (name in readCellData) or (nospName in readCellData):
                insertTable[f'{randCol}{randRow}'].value=readCellData + tableName
                thrownPoints[randRow, colIndex]=True
                foundInstances+=1

            if not isInMergedRange(readTable, readCell):
                if (name in readCellNextData) or (nospName in readCellNextData):
                    insertTable[f'{randCol}{randRow+1}'].value=readCellNextData + tableName
                    thrownPoints[randRow, colIndex]=True
                    foundInstances+=1
                else:
                    insertTable[f'{randCol}{randRow+1}'].value='#'

        except Exception as e:
             print(f'{colors.FAIL}Error: {e}.{colors.ENDC}')
        
    return foundInstances
#!---------------------------------------------------------


#?-------PSEUDO MAIN-------------------
#choosing data file:
data_files=getExcellFilenames()
group_tables=getPathsList()

os.makedirs('timetable/teacher', exist_ok=True)

with open('timetable/teacher/teacher_names.txt', 'r') as prof_names:
    for row in prof_names:
        saveFile=False
        name=row.split('|')[0]
        instanceCap=int(row.split('|')[1])
        instances=int(0)
        pdf_count=0

        if instanceCap in (0, '0'):
            continue
        
        print(f'{colors.WARNING}Prof:{name}{colors.ENDC}')

        writingBook=openpyxl.Workbook()
        insertTable=writingBook.active

        setTimeIntervals(insertTable)
        # applyDefaultMergeStyle(insertTable)
        setFontStyles(insertTable, 18)
        centerTableAlignment(insertTable)
        setTableDimensions(insertTable, 20, 40)


        for groupTable in group_tables:
            #open excel file
            openTable=load_workbook(groupTable)
            table=openTable.active

            nameInstances=countInstancesInText(name, extractAllTextFromExcel(groupTable))
            
            #if instances found
            if nameInstances != 0:
                print(f'processing {groupTable}# {nameInstances}')
                #monte carlo in file
                groupName=groupTable.split('/')[3].split('.')[0]
                transferProfClasses(name, nameInstances, table, insertTable, groupName)
                saveFile=True

            if nameInstances==instanceCap:
                break
        
        #save file
        if saveFile:
            applyDefaultMergeStyle(insertTable)
            print(f'{colors.OKGREEN}saved {name}.xlsx{colors.ENDC}')
            saveScheduleTable(writingBook, name)
#?-------------------------------------

