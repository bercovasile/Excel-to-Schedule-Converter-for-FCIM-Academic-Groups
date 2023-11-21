import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell

import os
import re

debug=False

class colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

#?-------Style functions------------------
def setTableDimensions(writingSheet, columnWidth, rowHeight):
    startColumn='A' #go +1 in loop #!  limit to 'G'
    endColumn='G'

    for row in range(0, 17): 
        writingSheet.row_dimensions[row].height=rowHeight

    for column in range(ord(startColumn), ord(endColumn)+1): 
        writingSheet.column_dimensions[chr(column)].width=columnWidth

def centerTableAlignment(writingSheet):
    startColumn='A'
    endColumn='G'

    for row in range(1, 17): #start at row 1, end at 16, increment by 2
        for column in range(ord(startColumn), ord(endColumn)+1):
            writingSheet[chr(column)+str(row)].alignment=Alignment(horizontal='center', vertical='center')   

def setFontStyles(writingSheet, fontSize):
    for row in range(1, 17, 2):
        writingSheet[f'A{row}'].font=Font(size=fontSize, bold=True)

    for column in range(ord('A'), ord('G')+1):
        writingSheet[f'{chr(column)}1'].font=Font(size=fontSize, bold=True)

def setTimeIntervals(writingSheet):
    #populate writeSheet with time intervals
    dayList=["Luni", "Marți", "Miercuri", "Joi", "Vineri", "Sâmbătă"]
    hourList=["8.00-9.30", "9.45-11.15", "11.30-13.00", "13.30-15.00", "15.15-16.45", "17.00-18.30", "18.45-20.15"]

    #populate writeSheet with time intervals
    startRow=3
    for timeInterval in hourList:
        writingSheet.cell(row=startRow, column=1, value=timeInterval)
        startRow+=2        

    #populate writeSheet with weekdays
    startColumn=2
    for day in dayList:
        writingSheet.cell(row=1, column=startColumn, value=day)
        startColumn+=1

def applyDefaultMergeStyle(writingSheet):
    startColumn='A' #go +1 in loop #!  limit to 'G'
    endColumn='G'

    for row in range(1, 17, 2): #start at row 1, end at 16, increment by 2
        for column in range(ord(startColumn), ord(endColumn)+1):
            currentCell=writingSheet[chr(column)+str(row)]
            underCell=writingSheet[chr(column)+str(row+1)]

            writingSheet.merge_cells(f'{currentCell.coordinate}:{underCell.coordinate}')

#?----------------------------------------

#?-------Helper functions for data--------
#?-------extraction-----------------------
def hasNoUpperBorder(cell):
    return cell.border.top.style in (None, "", "none")

def hasBottomBorder(cell, schedule):
    strCell=(str(cell).split('.')[1])[:-1]

    pattern = r'([A-Za-z]+)(\d+)'
    coord=re.match(pattern, strCell)

    nextCell=schedule[f'{coord.group(1)}{int(coord.group(2))+1}']
    
    if cell.border.bottom.style not in (None, "", "none") and nextCell.border.bottom.style not in (None, "", "none"):
        return False

    return cell.border.bottom.style not in (None, "", "none")

def getTimeInterval(schedule, row, cellLength):
    while schedule[f'C{row}'].value is None:
        row-=1
    else:
        if cellLength==12:
            return schedule[f'C{row-6}'].value
        else:
            return schedule[f'C{row}'].value

def isEven(schedule, row):
    cellCounter=1
    while schedule[f'C{row}'].value is None:
        row-=1
        cellCounter+=1
    else:
        if cellCounter==3:  
            return False
        elif cellCounter==6:
            return True

def isMerged(writingSheet, range):
    # Split the check_range into individual cell references
    cell_references = range.split(':')
    start_cell, end_cell = cell_references[0], cell_references[1]

        # Get the merged cell ranges in the worksheet
    merged_cells = writingSheet.merged_cells.ranges

        # Check if the specified range is merged
    is_merged = False

    for merged_range in merged_cells:
        if start_cell in merged_range and end_cell in merged_range:
            is_merged = True
            break

    return is_merged

def getCellLength(schedule, row, column):
    cellLength=1
    while hasNoUpperBorder(schedule[f'{column}{row}']):
        row-=1
        cellLength+=1
    else:
        return cellLength

def isInMergedRange(schedule, cell):
    if cell.coordinate in schedule.merged_cells:
        return True
    else:
        return False

def getFirstCellInRange(schedule, cell):
    for range in schedule.merged_cells.ranges:
        if cell.coordinate in range:
            return str(range).split(':')[0]

def getLastCellInRange(schedule, cell):
    for range in schedule.merged_cells.ranges:
        if cell.coordinate in range:
            return str(range).split(':')[1] 

def isMergedHorizontally(schedule, cell):
    firstCell=getFirstCellInRange(schedule, cell)
    lastCell=getLastCellInRange(schedule, cell)

    pattern = r'([A-Za-z]+)(\d+)'
    firstCoord=re.match(pattern, str(firstCell))
    secondCoord=re.match(pattern, str(lastCell))

    if(firstCoord.group(1)!=secondCoord.group(1)):
        return True
    else:
        return False

#?----------------------------------------

#?-------Functional functions-------------
def removeDuplicates(text):
    lines = text.split('\n')
    # Remove duplicates while preserving the order of the lines
    unique_lines = []
    seen_lines = set()

    for line in lines:
        if line not in seen_lines:
            seen_lines.add(line)
            unique_lines.append(line)

    # Join the unique lines back into a single string with newlines
    result_text = '\n'.join(unique_lines)
    return result_text

def findStartRow(schedule):
    for row in range(1, 15):
        if schedule[f'C{row}'].value == "8.00-9.30":
            return row
    else:
        return 0

def findGroupColumn(schedule, group, groupsRow):
    #extracting group names from excel file
    for column in schedule.iter_cols(min_row=groupsRow, max_row=groupsRow):
        for cell in column:
            if cell.value == group:
                columnLetter=openpyxl.utils.get_column_letter(cell.column)
                return columnLetter
        else:
            continue

    return None

def saveScheduleTable(writingBook, fileName, yearFile):
    if fileName is None:
        writingBook.close()
        return
    
    if '/' in fileName:
        fileName=fileName.replace('/', '-')

    filePath=f"group_schedules/{yearFile}/{fileName}.xlsx"

    if os.path.exists(filePath):
        os.remove(filePath)
    
    writingBook.save(filePath)
    writingBook.close()

def approximateCellLength(cellLength):
    if cellLength <= 3:
        return 3
    elif cellLength>3 and cellLength <= 7:
        return 6
    elif cellLength > 7:
        return 12
    return 0

def insertInTable(writingBook, insertColumn, timeInterval, cellLength, data, isEven):
    writingSheet=writingBook.active
    insertRow=0

    if timeInterval==None:
        print(f"{colors.FAIL}InsertInTable(): time interval is None.\033[0m")
        print(f"for data:{data}\nlen: {cellLength}{colors.ENDC}")
        return
        
    if cellLength not in (3, 6, 12):
        print(f"{colors.FAIL}Insert(): invalid cell length: {cellLength}.{colors.ENDC}")
        return

    for timeRow in range(3, 16, 2):
        if timeInterval==writingSheet[f'A{timeRow}'].value:
            insertRow=timeRow
            break

    if insertRow==0:
        print(f"\033[31mInsertInTable(): time interval not found: {timeInterval}\033[0m")
        return

    if cellLength==3:
        # print('Insert():small')
        insertColumn_char=openpyxl.utils.get_column_letter(insertColumn)
    
        mergedRange=f'{insertColumn_char}{insertRow}:{insertColumn_char}{insertRow+1}'
        if isMerged(writingSheet, mergedRange):
            writingSheet.unmerge_cells(f'{insertColumn_char}{insertRow}:{insertColumn_char}{insertRow+1}')
        
        if isEven:
            writingSheet.cell(row=insertRow+1, column=insertColumn).value=data
        else:
            writingSheet.cell(row=insertRow, column=insertColumn).value=data
        
        return
    
    elif cellLength==6:
        # print('Insert():regular')
        writingSheet.cell(row=insertRow, column=insertColumn).value=data

    elif cellLength==12:
        # print('Insert():double')
        insertColumn_char=openpyxl.utils.get_column_letter(insertColumn)
        writingSheet.unmerge_cells(f'{insertColumn_char}{insertRow}:{insertColumn_char}{insertRow+1}')
        writingSheet.unmerge_cells(f'{insertColumn_char}{insertRow+2}:{insertColumn_char}{insertRow+3}')
        writingSheet.merge_cells(f'{insertColumn_char}{insertRow}:{insertColumn_char}{insertRow+3}')
        writingSheet.cell(row=insertRow, column=insertColumn).value=data

def extractAndTransferToTable(writingBook, readingBook, columnIndex):
    tempData=""
    rowNumber=findStartRow(schedule)

    cellLength=0
    dayCounter=-1
    dayIndex=2
    lastInsertLength=6
    

    while True:
        currentCell_reference=f'{columnIndex}{rowNumber}'
        currentCell_value=readingBook[currentCell_reference].value
        currentCell=readingBook[currentCell_reference]


        dayCounter+=1
        cellLength+=1

        if cellLength > 12:
            print(f'{colors.FAIL}Extract(): Cell length > 12.{colors.ENDC}')

        if debug:
            if currentCell_value is None:
                print(f'{colors.WARNING}{currentCell_reference}|{colors.ENDC}')
            else:
                print(f'{currentCell_reference}={currentCell_value}')
                #print(f'{colors.BOLD}->len:{cellLength}{colors.ENDC}')

        if currentCell_value is None:
            if isInMergedRange(readingBook, currentCell):
                if isMergedHorizontally(readingBook, currentCell):
                    firstInRange=getFirstCellInRange(readingBook, currentCell)
                    if readingBook[firstInRange].value is not None:
                        cellData=readingBook[firstInRange].value
                        if cellData not in tempData:
                            tempData+=cellData + '\n'

        elif currentCell_value != "MCE":
            tempData+=str(currentCell_value)+'\n'
            
        if debug:
            if hasBottomBorder(currentCell, readingBook):
                print(f"{colors.OKBLUE}#MET_BRDR{colors.ENDC}")
            elif tempData.count('\n') == 3 and cellLength in (3, 6):
                print(f"{colors.OKBLUE}#LINELEN=3_LEN=(3,6){colors.ENDC}")
            elif cellLength==6 and tempData=="":
                print(f"{colors.OKBLUE}#NODATA_LEN=6{colors.ENDC}")
            elif cellLength==12 and tempData!="":
                print(f"{colors.OKBLUE}#LEN=12_DATA{colors.ENDC}")

        if hasBottomBorder(currentCell, readingBook) or (cellLength==12 and tempData!="") or (tempData.count('\n') == 3 and cellLength in (3, 6)) or (cellLength==6 and tempData==""):
            if tempData!="":
                if debug:
                    print(f'{colors.HEADER}----------------start_insert{colors.ENDC}')
                    print(f'{colors.BOLD}{tempData}ln:{cellLength}{colors.ENDC}')

                if lastInsertLength==3 and cellLength==9:
                    cellLength=6

                timeInterval=getTimeInterval(readingBook, rowNumber, cellLength)
                isEvenCell=isEven(readingBook, rowNumber)
                insertInTable(writingBook, dayIndex, timeInterval, cellLength, tempData, isEvenCell)

                lastInsertLength=cellLength
                
                if debug:
                    print(f'{colors.OKCYAN}----------------end_insert{colors.ENDC}')

            tempData=""
            cellLength=0

        if dayCounter == 42:
            if tempData!="":
                print(f"{colors.FAIL}Extract({currentCell_reference}): Data out of day bounds.{colors.ENDC}")

            tempData=""
            cellLength=0
            dayIndex+=1
            dayCounter=-1
    
        rowNumber+=1
        if rowNumber==260:
            break

def getExcellFilenames():
    excelFiles=[]
    for filename in os.listdir():
        if '.xlsx' in filename and '$' not in filename:
            excelFiles.append(filename)
    
    excelFiles[0],excelFiles[2],excelFiles[3]=excelFiles[3],excelFiles[0], excelFiles[2]
    return excelFiles

def getGroupNames(schedule):
    #lista grupe dupa nume
    fcimGroups=[]
    
    dataRow=""
    #get group start row
    for row in range(1, 20):
        #print(f"B{row}=={schedule[f'B{row}'].value}")
        if schedule[f'B{row}'].value =='Grupele':
            dataRow=row
            break
         
    #popularea listei cu numele grupelor din excel
    for cell in schedule.iter_cols(min_row=dataRow, max_row=dataRow, min_col=5):
        if cell[0].value is not None :
            fcimGroups.append(cell[0].value)
        else:
            break

    return fcimGroups, dataRow

#?----------------------------------------

#?-------Setting stuff up-----------------
#choosing data file:
#!get it from the server
# data_file='Anul_III_2023_Semestrul_V.xlsx'
# yearFile='anul_III'

# #load the data workbook
# readingBook=load_workbook(data_file)
#     #select datasheet
# schedule=readingBook.active

# #creating writing book
# writingBook=openpyxl.Workbook()
# writingSheet=writingBook.active
# #?-----------------------------------------


# #?------Style writing file-----------------
# setTimeIntervals(writingSheet)
# applyDefaultMergeStyle(writingSheet)
# # ?-----------------------------------------

# fcimGroups, groupRow=getGroupNames(schedule)
# print(fcimGroups)

#*---Searching column letter and saving----
#*---custom named file.--------------------

# searchValue=input("Search group: ")
# searchResult_columnLetter=findGroupColumn(schedule, searchValue, groupRow)

# if searchResult_columnLetter is not None:
#     print(f"found {searchValue} on column {searchResult_columnLetter}.")
#     extractAndTransferToTable(writingBook, schedule, searchResult_columnLetter)

#     setTableDimensions(writingSheet, 20, 40)
#     centerTableAlignment(writingSheet)
#     setFontStyles(writingSheet, 18)

#     saveScheduleTable(writingBook, searchValue, yearFile)
# else:
#     print(f" {searchValue} not found.")

excelFilenames=getExcellFilenames()
excelFileCount=0

yearNames=['anul_I', 'anul_II','anul_III', 'anul_IV']
yearFileCount=0

for excelFile in excelFilenames:
    print(f'{colors.BOLD}{colors.OKGREEN}>>>>>>>>{yearNames[yearFileCount]}{colors.ENDC}')
    readingBook=load_workbook(excelFile)
    schedule=readingBook.active 

    fcimGroups, groupRow=getGroupNames(schedule)
    
    for group in fcimGroups:
        searchResult_columnLetter=findGroupColumn(schedule, group, groupRow)
        
        writingBook=openpyxl.Workbook()
        writingSheet=writingBook.active
        #apply styles
        setTimeIntervals(writingSheet)
        applyDefaultMergeStyle(writingSheet)

        if searchResult_columnLetter is not None:
            print(f"found {group} on column {searchResult_columnLetter}.")
            extractAndTransferToTable(writingBook, schedule, searchResult_columnLetter)

            setTableDimensions(writingSheet, 20, 40)
            centerTableAlignment(writingSheet)
            setFontStyles(writingSheet, 18)

            saveScheduleTable(writingBook, group, yearNames[yearFileCount])
        else:
            print(f" {group} not found.")
    yearFileCount+=1
    excelFileCount+=1


# for group in fcimGroups:
#         #find column letter 
#         searchResult_columnLetter=findGroupColumn(schedule, group, groupRow)

#         #create a new writebook
#         writingBook=openpyxl.Workbook()
#         writingSheet=writingBook.active
#         #apply styles
#         setTimeIntervals(writingSheet)
#         applyDefaultMergeStyle(writingSheet)

#         if searchResult_columnLetter is not None:
#             print(f"found {group} on column {searchResult_columnLetter}.")
#             extractAndTransferToTable(writingBook, schedule, searchResult_columnLetter)

#             setTableDimensions(writingSheet, 20, 40)
#             centerTableAlignment(writingSheet)
#             setFontStyles(writingSheet, 18)

#             saveScheduleTable(writingBook, group, yearFile)
#         else:
#             print(f" {group} not found.")
#*-----------------------------------------

readingBook.close()

#!TAKE A LOOK AT
#*ANUL 2:
# TI-224
# TI-225
# TI-227

#*ANUL 3:
#FAF-221
